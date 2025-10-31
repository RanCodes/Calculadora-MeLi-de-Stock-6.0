import io
from math import isclose
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from data_processor import (
    leer_ml,
    leer_odoo,
    unir_y_validar,
    calcular,
    preparar_resultado_final,
    exportar_excel,
)
from utils import calcular_precio_publicacion_ml

def crear_datos_ejemplo():
    """
    Crea datos de ejemplo basados en la estructura real de los archivos.
    """
    # Datos de ejemplo MercadoLibre
    ml_data = {
        'ITEM_ID': ['MLA934071512', 'MLA864175834', 'MLA123456789', 'MLA987654321', 'MLA555666777'],
        'VARIATION_ID': ['94135189655', '175837150784', None, '123456789', None],
        'SKU': ['LED7012795', 'CORNPR06WW', 'TCL45310', 'MMM42385', 'NOEXISTE123'],
        'TITLE': [
            'Lampara Sodio 250w E40 Osram Alumbrado Público',
            'Macroled Panel Plafón Redondo Led 6w Cálido Black Npr06',
            'Módulo ciego BLANCO',
            '3M™ Cinta de Empaque 301 - 48mm x 50m',
            'Producto sin match en Odoo'
        ],
        'QUANTITY': [232, 106, 50, 200, 10],
        'PRICE': [21997.46, 8164.02, 250.00, 2500.00, 1000.00],
        'CURRENCY_ID': ['$', '$', '$', '$', '$'],
        'FEE_PER_SALE_MARKETPLACE_V2': [
            '14.50% + $1095.00',
            '12.00% + $800.00',
            '15.00% + $500.00',
            '13.50% + $750.00',
            '16.00% + $1200.00'
        ],
        'COST_OF_FINANCING_MARKETPLACE': ['4.00%', '3.50%', '0.00%', '5.00%', '4.50%'],
        'LISTING_TYPE_V3': ['gold_special', 'gold_pro', 'free', 'gold_special', 'gold_pro'],
        'SHIPPING_METHOD ': [
            'Mercado Envíos por mi cuenta',
            'Mercado Envíos Clásico',
            'Mercado envíos POR MI CUENTA (Flex)',
            None,
            'Retiro en tienda'
        ]
    }

    # Datos de ejemplo Odoo
    odoo_data = {
        'Código Neored': ['LED7012795', 'CORNPR06WW', 'TCL45310', 'MMM42385', 'EXTRA12345'],
        'Nombre': [
            'Lámpara Sodio 250W E40 Osram',
            'Panel LED Redondo 6W Cálido',
            '1/2 Módulo ciego BLANCO',
            '3M™ Cinta de Empaque 301 - 48mm x 50m',
            'Producto extra sin match en ML'
        ],
        'Cantidad a mano': [250, 120, 60, 576, 100],
        'Precio Tarifa': [18500.00, 6800.00, 184.05, 1915.72, 2000.00],
        'Impuestos del cliente': ['IVA Ventas 21%', 'IVA Ventas 21%', 'IVA Ventas 21%', 'IVA Ventas 21%', 'IVA Ventas 21%']
    }

    df_ml = pd.DataFrame(ml_data)
    df_odoo = pd.DataFrame(odoo_data)
    return df_ml, df_odoo

def preparar_df_para_calculo():
    """Prepara un DataFrame unido listo para pasar a ``calcular``."""
    df_ml, df_odoo = crear_datos_ejemplo()
    from utils import parse_fee_combo, parse_pct, extract_tax_percentage

    df_ml['fee_pct'], df_ml['fee_fixed'] = zip(*df_ml['FEE_PER_SALE_MARKETPLACE_V2'].apply(parse_fee_combo))
    df_ml['financing_pct'] = df_ml['COST_OF_FINANCING_MARKETPLACE'].apply(parse_pct)
    df_odoo['tax_pct'] = df_odoo['Impuestos del cliente'].apply(extract_tax_percentage)

    return unir_y_validar(df_ml, df_odoo)


def _build_ml_excel_bytes(rows):
    headers = [
        'ITEM_ID',
        'VARIATION_ID',
        'SKU',
        'TITLE',
        'QUANTITY',
        'PRICE',
        'CURRENCY_ID',
        'FEE_PER_SALE_MARKETPLACE_V2',
        'COST_OF_FINANCING_MARKETPLACE',
        'LISTING_TYPE_V3',
        'SHIPPING_METHOD ',
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja1'
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def test_leer_ml_detects_formulas_and_converts_to_values():
    formula_rows = [
        [
            'MLA123',
            None,
            'SKU123',
            'Producto de prueba',
            '=@[qty]',
            '=@[price]',
            '$',
            '=@[fee] + 0',
            '=@[financing] + 0',
            'gold_special',
            'Mercado Envíos por mi cuenta',
        ]
    ]
    value_rows = [
        [
            'MLA123',
            None,
            'SKU123',
            'Producto de prueba',
            25,
            1999.99,
            '$',
            '14.50% + $1095.00',
            '4.00%',
            'gold_special',
            'Mercado Envíos por mi cuenta',
        ]
    ]

    formula_bytes = _build_ml_excel_bytes(formula_rows)
    values_bytes = _build_ml_excel_bytes(value_rows)

    with patch('data_processor._convert_ml_workbook_to_values', return_value=values_bytes) as mocked:
        df_result = leer_ml(io.BytesIO(formula_bytes))

    assert mocked.called, "Se esperaba que se convirtiera el workbook a valores."
    assert not df_result.empty
    assert df_result.loc[0, 'SKU'] == 'SKU123'
    assert df_result.loc[0, 'PRICE'] == 1999.99
    assert df_result.loc[0, 'QUANTITY'] == 25

def test_procesamiento():
    """
    Prueba el procesamiento completo con datos de ejemplo.
    """
    print("🧪 Iniciando prueba con datos de ejemplo")
    print("=" * 50)

    df_ml, df_odoo = crear_datos_ejemplo()
    print("📊 Datos de entrada creados:")
    print(f"- MercadoLibre: {len(df_ml)} items")
    print(f"- Odoo: {len(df_odoo)} productos")
    print("\n🔄 Simulando procesamiento...")
    from utils import parse_fee_combo, parse_pct
    df_ml['fee_pct'], df_ml['fee_fixed'] = zip(*df_ml['FEE_PER_SALE_MARKETPLACE_V2'].apply(parse_fee_combo))
    df_ml['financing_pct'] = df_ml['COST_OF_FINANCING_MARKETPLACE'].apply(parse_pct)
    from utils import extract_tax_percentage
    df_odoo['tax_pct'] = df_odoo['Impuestos del cliente'].apply(extract_tax_percentage)
    print("✅ Campos parseados correctamente")
    print("\n🔗 Uniendo datos por SKU...")
    df_merged = unir_y_validar(df_ml, df_odoo)
    total_items = len(df_merged)
    matched_items = len(df_merged[df_merged['Código Neored'].notna()])
    print(f"- Total items: {total_items}")
    print(f"- Matches encontrados: {matched_items}")
    print(f"- Tasa de match: {matched_items/total_items*100:.1f}%")
    no_match = df_merged[df_merged['Código Neored'].isna()][['SKU', 'TITLE']]
    if not no_match.empty:
        print("\n⚠️ Items sin match:")
        for idx, row in no_match.iterrows():
            print(f"  - {row['SKU']}: {row['TITLE'][:50]}...")
    print("\n💰 Calculando precios (base: tarifa, sin impuestos)...")
    df_calculated = calcular(df_merged, base_financiacion='tarifa', incluir_impuestos=False)
    assert 'IVA' in df_calculated.columns

    producto_led = df_calculated[df_calculated['SKU'] == 'LED7012795'].iloc[0]
    tarifa_led = float(df_odoo.loc[df_odoo['Código Neored'] == 'LED7012795', 'Precio Tarifa'].iloc[0])
    fee_pct_led = float(df_ml.loc[df_ml['SKU'] == 'LED7012795', 'fee_pct'].iloc[0])
    fee_fixed_led = float(df_ml.loc[df_ml['SKU'] == 'LED7012795', 'fee_fixed'].iloc[0])
    financing_pct_led = float(df_ml.loc[df_ml['SKU'] == 'LED7012795', 'financing_pct'].iloc[0])
    tax_pct_led = float(df_odoo.loc[df_odoo['Código Neored'] == 'LED7012795', 'tax_pct'].iloc[0])

    (
        precio_led,
        cargo_led,
        cuotas_led,
        ret_led,
        recibis_led,
        denominador_invalido,
    ) = calcular_precio_publicacion_ml(
        tarifa_neta=tarifa_led,
        porcentaje_comision=fee_pct_led,
        porcentaje_financiacion=financing_pct_led,
        porcentaje_retenciones=0.0,
        costo_fijo=fee_fixed_led,
    )
    assert not denominador_invalido

    iva_led = precio_led * tax_pct_led / (1 + tax_pct_led)
    recargo_pct_led = precio_led * fee_pct_led

    assert isclose(producto_led['Precio final'], precio_led, rel_tol=1e-04)
    assert isclose(producto_led['Cargo por vender ($)'], cargo_led, rel_tol=1e-04)
    assert isclose(producto_led['Recargo financiación (importe)'], cuotas_led, rel_tol=1e-04)
    assert isclose(producto_led['Retenciones ML ($)'], ret_led, rel_tol=1e-04)
    assert isclose(producto_led['Recibis ($)'], recibis_led, rel_tol=1e-04)
    assert isclose(producto_led['Recargo % ML (importe)'], recargo_pct_led, rel_tol=1e-04)
    assert isclose(producto_led['Recargo fijo ML ($)'], fee_fixed_led, rel_tol=1e-04)
    assert isclose(producto_led['IVA'], iva_led, rel_tol=1e-04)
    print("\n📋 Preparando resultado final...")
    df_resultado = preparar_resultado_final(df_calculated, incluir_impuestos=False)
    assert 'IVA' in df_resultado.columns
    assert '% Stock' in df_resultado.columns
    resultado_led = df_resultado[df_resultado['SKU'] == 'LED7012795'].iloc[0]
    assert isclose(resultado_led['Cargo por vender ($)'], cargo_led, rel_tol=1e-04)
    assert isclose(resultado_led['Retenciones ML ($)'], ret_led, rel_tol=1e-04)
    assert isclose(resultado_led['Recibis ($)'], recibis_led, rel_tol=1e-04)
    assert isclose(resultado_led['IVA'], iva_led, rel_tol=1e-04)
    assert isclose(resultado_led['Precio final'], precio_led, rel_tol=1e-04)
    assert resultado_led['% Stock'] == resultado_led['Stock']
    print("✅ Procesamiento completado")
    print("\n📊 RESULTADOS DETALLADOS:")
    print("=" * 50)
    for idx, row in df_resultado.iterrows():
        if row['Precio final'] > 0:
            print(f"\n🏷️ SKU: {row['SKU']}")
            print(f"📦 Producto: {row['Descripción del producto'][:60]}")
            print(f"📊 Stock: {row['Stock']} unidades")
            print(f"💵 Precio Tarifa: ${row['Precio de Tarifa']:,.2f}")
            print(f"🎯 Precio Final: ${row['Precio final']:,.2f}")
            print("📈 Desglose:")
            print(f"   - Recargo % ML ({row['% ML aplicado']:.1f}%): ${row['Recargo % ML (importe)']:,.2f}")
            print(f"   - Recargo fijo ML: ${row['Recargo fijo ML ($)']:,.2f}")
            print(f"   - Recargo financiación ({row['% financiación aplicado']:.1f}%): ${row['Recargo financiación (importe)']:,.2f}")
            print(f"   - Retenciones ML: ${row['Retenciones ML ($)']:,.2f}")
            print(f"   - Recibís neto: ${row['Recibis ($)']:,.2f}")
            if 'Recargo envío ($)' in df_resultado.columns:
                print(f"   - Recargo envío: ${row.get('Recargo envío ($)', 0):,.2f}")
            print(f"📋 Tipo: {row['Tipo de publicación']}")
            if row['Notas/Flags']:
                print(f"⚠️ Advertencias: {row['Notas/Flags']}")
    print("\n" + "=" * 50)
    print("🧪 PRUEBA CON IMPUESTOS INCLUIDOS")
    print("=" * 50)
    df_calculated_tax = calcular(df_merged, base_financiacion='tarifa', incluir_impuestos=True)
    assert 'IVA' in df_calculated_tax.columns
    producto_led_tax = df_calculated_tax[df_calculated_tax['SKU'] == 'LED7012795'].iloc[0]
    (
        precio_led_tax,
        cargo_led_tax,
        cuotas_led_tax,
        ret_led_tax,
        recibis_led_tax,
        denominador_invalido_tax,
    ) = calcular_precio_publicacion_ml(
        tarifa_neta=tarifa_led * (1 + tax_pct_led),
        porcentaje_comision=fee_pct_led,
        porcentaje_financiacion=financing_pct_led,
        porcentaje_retenciones=0.0,
        costo_fijo=fee_fixed_led,
    )
    assert not denominador_invalido_tax

    iva_led_tax = precio_led_tax * tax_pct_led / (1 + tax_pct_led)
    recargo_pct_led_tax = precio_led_tax * fee_pct_led

    assert isclose(producto_led_tax['Precio final'], precio_led_tax, rel_tol=1e-04)
    assert isclose(producto_led_tax['Cargo por vender ($)'], cargo_led_tax, rel_tol=1e-04)
    assert isclose(producto_led_tax['Recargo financiación (importe)'], cuotas_led_tax, rel_tol=1e-04)
    assert isclose(producto_led_tax['Retenciones ML ($)'], ret_led_tax, rel_tol=1e-04)
    assert isclose(producto_led_tax['Recibis ($)'], recibis_led_tax, rel_tol=1e-04)
    assert isclose(producto_led_tax['Recargo % ML (importe)'], recargo_pct_led_tax, rel_tol=1e-04)
    assert isclose(producto_led_tax['IVA'], iva_led_tax, rel_tol=1e-04)
    df_resultado_tax = preparar_resultado_final(df_calculated_tax, incluir_impuestos=True)
    assert 'IVA' in df_resultado_tax.columns
    resultado_led_tax = df_resultado_tax[df_resultado_tax['SKU'] == 'LED7012795'].iloc[0]
    assert isclose(resultado_led_tax['Cargo por vender ($)'], cargo_led_tax, rel_tol=1e-04)
    assert isclose(resultado_led_tax['Retenciones ML ($)'], ret_led_tax, rel_tol=1e-04)
    assert isclose(resultado_led_tax['Recibis ($)'], recibis_led_tax, rel_tol=1e-04)
    assert isclose(resultado_led_tax['IVA'], iva_led_tax, rel_tol=1e-04)
    assert isclose(resultado_led_tax['Precio final'], precio_led_tax, rel_tol=1e-04)
    df_resultado_stock_50 = preparar_resultado_final(
        df_calculated,
        incluir_impuestos=False,
        porcentaje_stock=50,
    )
    expected_stock_50 = (
        df_calculated['Cantidad a mano'].fillna(0) * 0.5
    ).round().astype(int).reset_index(drop=True)
    pd.testing.assert_series_equal(
        df_resultado_stock_50['% Stock'].reset_index(drop=True),
        expected_stock_50,
    )
    print("\n📊 Comparación con y sin impuestos (primeros 3 items):")
    comparacion_cols = ['SKU', 'Precio de Tarifa', 'Tarifa + impuestos', 'Precio final']
    df_comp = df_resultado_tax[df_resultado_tax['Precio final'] > 0][comparacion_cols].head(3)
    for idx, row in df_comp.iterrows():
        print(f"\n🏷️ SKU: {row['SKU']}")
        print(f"💵 Tarifa base: ${row['Precio de Tarifa']:,.2f}")
        print(f"💰 Tarifa + IVA: ${row['Tarifa + impuestos']:,.2f}")
        print(f"🎯 Precio final: ${row['Precio final']:,.2f}")
        aumento = ((row['Precio final'] / row['Precio de Tarifa']) - 1) * 100
        print(f"📈 Aumento total: {aumento:.1f}%")
    print(f"\n💾 Generando archivo Excel de ejemplo...")
    excel_bytes = exportar_excel(df_resultado)
    with open('ML_precios_y_stock_calculados_EJEMPLO.xlsx', 'wb') as f:
        f.write(excel_bytes)
    wb = load_workbook(filename=io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert 'IVA' in headers
    print("✅ Archivo generado: ML_precios_y_stock_calculados_EJEMPLO.xlsx")
    print(f"\n📈 RESUMEN ESTADÍSTICO:")
    print("=" * 30)
    items_validos = df_resultado[df_resultado['Precio final'] > 0]
    if not items_validos.empty:
        print(f"Items procesados: {len(items_validos)}")
        print(f"Precio promedio: ${items_validos['Precio final'].mean():,.2f}")
        print(f"Precio mínimo: ${items_validos['Precio final'].min():,.2f}")
        print(f"Precio máximo: ${items_validos['Precio final'].max():,.2f}")
        recargo_total = (
            items_validos['Recargo % ML (importe)'] +
            items_validos['Recargo fijo ML ($)'] +
            items_validos['Recargo financiación (importe)']
        )
        if 'Recargo envío ($)' in items_validos.columns:
            recargo_total = recargo_total + items_validos['Recargo envío ($)']
        print(f"\nRecargo promedio ML: ${recargo_total.mean():,.2f}")
        porcentaje_recargo = (recargo_total / items_validos['Precio de Tarifa']).mean() * 100
        print(f"Recargo promedio %: {porcentaje_recargo:.1f}%")
    print("\n🎉 ¡Prueba completada exitosamente!")
    return df_resultado


def test_calcular_precio_publicacion_ml():
    """Valida el cálculo directo del precio de publicación."""

    tarifa_neta = 12513.10
    porcentaje_comision = 0.145
    porcentaje_financiacion = 0.04
    porcentaje_retenciones = 0.01
    costo_fijo = 2190.0

    (
        precio_publicacion,
        cargo_por_vender,
        recargo_financiacion,
        retenciones,
        recibis,
        denominador_invalido,
    ) = calcular_precio_publicacion_ml(
        tarifa_neta=tarifa_neta,
        porcentaje_comision=porcentaje_comision,
        porcentaje_financiacion=porcentaje_financiacion,
        porcentaje_retenciones=porcentaje_retenciones,
        costo_fijo=costo_fijo,
    )

    assert not denominador_invalido
    assert isclose(precio_publicacion, 18264.72, rel_tol=1e-04)
    assert isclose(recibis, tarifa_neta, rel_tol=1e-06)
    # Validar que los componentes sumen correctamente
    total_descuentos = cargo_por_vender + recargo_financiacion + retenciones
    assert isclose(precio_publicacion - total_descuentos, tarifa_neta, rel_tol=1e-06)

def test_recargo_envio_fijo_aplica_solo_a_envios_por_cuenta_propia():
    df_merged = preparar_df_para_calculo()
    df_calculado = calcular(df_merged, tipo_recargo_envio='Fijo ($)', valor_recargo_envio=150)

    skus_con_recargo = set(df_calculado.loc[df_calculado['Recargo envío ($)'] > 0, 'SKU'])
    skus_esperados = {'LED7012795', 'TCL45310'}
    assert skus_con_recargo == skus_esperados

    filas_sin_recargo = df_calculado[~df_calculado['SKU'].isin(skus_esperados)]['Recargo envío ($)']
    assert all(isclose(valor, 0.0, abs_tol=1e-9) for valor in filas_sin_recargo)

def test_recargo_envio_porcentaje_respeta_mascara():
    df_merged = preparar_df_para_calculo()
    df_calculado = calcular(df_merged, tipo_recargo_envio='Porcentaje (%)', valor_recargo_envio=10)

    recargo_led = df_calculado.loc[df_calculado['SKU'] == 'LED7012795', 'Recargo envío ($)'].iloc[0]
    recargo_modulo = df_calculado.loc[df_calculado['SKU'] == 'TCL45310', 'Recargo envío ($)'].iloc[0]
    assert isclose(recargo_led, 1850.0, rel_tol=1e-4)
    assert isclose(recargo_modulo, 18.41, rel_tol=1e-4)

    otros = df_calculado[~df_calculado['SKU'].isin({'LED7012795', 'TCL45310'})]['Recargo envío ($)']
    assert all(isclose(valor, 0.0, abs_tol=1e-9) for valor in otros)

def test_recargo_envio_sin_columna_shipping_no_aplica():
    df_merged = preparar_df_para_calculo()
    df_sin_shipping = df_merged.drop(columns=['SHIPPING_METHOD '])
    assert 'SHIPPING_METHOD ' not in df_sin_shipping.columns

    df_calculado = calcular(df_sin_shipping, tipo_recargo_envio='Fijo ($)', valor_recargo_envio=200)
    assert all(isclose(valor, 0.0, abs_tol=1e-9) for valor in df_calculado['Recargo envío ($)'])

def test_parseo_individual():
    """
    Prueba las funciones de parseo individualmente.
    """
    from utils import parse_money, parse_pct, parse_fee_combo
    print("\n🔧 PRUEBA DE FUNCIONES DE PARSEO:")
    print("=" * 40)
    print("\n💰 Pruebas parse_money:")
    money_tests = [
        "$1,095.00", "1095", "1.095,50", "$2.500,75", "0", "", None
    ]
    for test in money_tests:
        result = parse_money(test)
        print(f"  '{test}' -> {result}")
    print("\n📊 Pruebas parse_pct:")
    pct_tests = [
        "14.50%", "4.00%", "0.04", "4", "21%", "0", "", None
    ]
    for test in pct_tests:
        result = parse_pct(test)
        print(f"  '{test}' -> {result:.4f}")
    print("\n🔀 Pruebas parse_fee_combo:")
    combo_tests = [
        "14.50% + $1095.00",
        "12.00% + $800.00",
        "15.00%",
        "$500.00",
        "16% + $1,200.00",
        "",
        None
    ]
    for test in combo_tests:
        pct, fixed = parse_fee_combo(test)
        print(f"  '{test}' -> {pct:.4f}, ${fixed:.2f}")

if __name__ == "__main__":
    test_parseo_individual()
    test_procesamiento()
    print("\n✨ Todas las pruebas completadas. Revisa el archivo Excel generado!")